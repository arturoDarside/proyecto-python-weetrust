import imaplib
import email
from email.header import decode_header
import unicodedata
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import pandas as pd
from openpyxl import load_workbook
from scripts import config
import os
import re
import chardet
from pathlib import Path
from datetime import datetime
import logging

# Configuración básica
EXCEL_FILE = "solicitudes_estadia.xlsx"
BASE_FOLDER = Path("F:/Descargas/CVs")  # Carpeta base para guardar CVs
EMAIL_KEYWORDS = ["solicitud de estadia", "practicas", "estadia profesional"]

# Configurar logging para seguimiento
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s: %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger('EmailProcessor')

# Flag global para detener la lectura
detener_busqueda = False  

def normalizar_texto(texto):
    """Normaliza texto eliminando acentos y caracteres especiales"""
    if not texto:
        return ""
    if isinstance(texto, bytes):
        try:
            detected = chardet.detect(texto)
            texto = texto.decode(detected['encoding'] or 'utf-8', errors='ignore')
        except:
            texto = texto.decode('utf-8', errors='ignore')
    else:
        texto = str(texto)
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')
    return texto.strip()

def limpiar_nombre(nombre):
    """Limpia un nombre para usarlo como nombre de archivo válido"""
    nombre = unicodedata.normalize("NFD", nombre)
    nombre = nombre.encode("ascii", "ignore").decode("utf-8")
    nombre = re.sub(r'[^\w\-_. ]', '', nombre)
    return nombre.replace(" ", "_").lower()

def extraer_cuerpo(mensaje):
    """Extrae el cuerpo de texto de un mensaje de correo"""
    cuerpo = ""
    if mensaje.is_multipart():
        for parte in mensaje.walk():
            content_type = parte.get_content_type()
            content_disposition = str(parte.get("Content-Disposition"))
            if "attachment" not in content_disposition and content_type == "text/plain":
                cuerpo = parte.get_payload(decode=True)
                if cuerpo:
                    detected = chardet.detect(cuerpo)
                    cuerpo = cuerpo.decode(detected['encoding'] or 'utf-8', errors='ignore')
                    break
    else:
        cuerpo = mensaje.get_payload(decode=True)
        if cuerpo:
            detected = chardet.detect(cuerpo)
            cuerpo = cuerpo.decode(detected['encoding'] or 'utf-8', errors='ignore')
    return cuerpo.strip()

def extraer_datos_estudiante(cuerpo):
    """Extrae datos del estudiante del cuerpo del correo"""
    # Patrones de búsqueda mejorados con límites de línea
    nombre = re.search(r'nombre[:\-]?\s*(.*?)(?:\n|$)', cuerpo, re.IGNORECASE)
    carrera = re.search(r'carrera[:\-]?\s*(.*?)(?:\n|$)', cuerpo, re.IGNORECASE)
    universidad = re.search(r'universidad[:\-]?\s*(.*?)(?:\n|$)', cuerpo, re.IGNORECASE)
    telefono = re.search(r'telefono[:\-]?\s*(.*?)(?:\n|$)', cuerpo, re.IGNORECASE)
    
    return {
        "nombre": nombre.group(1).strip() if nombre else "",
        "carrera": carrera.group(1).strip() if carrera else "",
        "universidad": universidad.group(1).strip() if universidad else "",
        "telefono": telefono.group(1).strip() if telefono else ""
    }

def generar_id_y_carpeta():
    """Genera un ID único y carpeta para guardar CVs"""
    try:
        id_registro = 1
        
        # Revisar si el archivo Excel existe y calcular el próximo ID
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            max_rows = ws.max_row
            if max_rows > 1:  # Si hay datos (no solo encabezados)
                id_registro = max_rows  # El número de la fila es el ID
            wb.close()
        
        # Crear carpeta para guardar documentos con ese ID
        carpeta_id = BASE_FOLDER / f"ID_{id_registro}"
        carpeta_id.mkdir(parents=True, exist_ok=True)
        
        return id_registro, carpeta_id
    except Exception as e:
        logger.error(f"Error generando ID: {e}")
        # Carpeta de emergencia con timestamp si hay error
        fallback_folder = BASE_FOLDER / f"error_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        fallback_folder.mkdir(parents=True, exist_ok=True)
        return 0, fallback_folder

def guardar_cv_adjuntos(mensaje, id_registro, nombre_estudiante, carpeta_destino):
    """Guarda adjuntos del correo en la carpeta especificada"""
    archivos_guardados = []
    
    nombre_base = limpiar_nombre(nombre_estudiante or "sin_nombre")
    
    for parte in mensaje.walk():
        if parte.get_content_maintype() == 'multipart':
            continue
        if parte.get('Content-Disposition') is None:
            continue

        filename = parte.get_filename()
        if filename:
            try:
                # Decodificar nombre del archivo
                filename = decode_header(filename)[0][0]
                if isinstance(filename, bytes):
                    filename = normalizar_texto(filename)
                
                # Limpiar nombre de archivo
                filename = limpiar_nombre(filename)
                
                # Obtener extensión o usar .pdf por defecto
                extension = os.path.splitext(filename)[1] or ".pdf"
                
                # Crear nombre final con ID para relacionar con el Excel
                nombre_final = f"ID_{id_registro}_{nombre_base}{extension}"
                ruta_completa = carpeta_destino / nombre_final
                
                # Guardar archivo
                with open(ruta_completa, "wb") as f:
                    f.write(parte.get_payload(decode=True))
                archivos_guardados.append(str(ruta_completa))
                logger.info(f"📎 CV guardado: {ruta_completa}")
            except Exception as e:
                logger.error(f"⚠ Error al guardar adjunto: {e}")

    return archivos_guardados

def guardar_en_excel(id_registro, datos, ruta_cv):
    """Guarda los datos en Excel con más campos y formateando correctamente"""
    try:
        # Normalizar todos los datos
        nombre = normalizar_texto(datos.get("nombre", ""))
        carrera = normalizar_texto(datos.get("carrera", ""))
        universidad = normalizar_texto(datos.get("universidad", ""))
        remitente = normalizar_texto(datos.get("remitente", ""))
        telefono = normalizar_texto(datos.get("telefono", ""))
        ruta_cv = normalizar_texto(ruta_cv)
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Crear DataFrame con los datos y más columnas útiles
        df = pd.DataFrame([[
            id_registro, 
            nombre, 
            carrera, 
            universidad, 
            remitente,
            telefono,
            f"ID_{id_registro}", # Carpeta donde se guardan los archivos
            ruta_cv,
            fecha,
            "" # Campo de notas/observaciones
        ]], columns=[
            'ID', 
            'Nombre', 
            'Carrera', 
            'Universidad', 
            'Correo',
            'Teléfono', 
            'ID_Carpeta',
            'Ruta_CV', 
            'Fecha_Recepción',
            'Observaciones'
        ])

        # Si no existe el archivo, creamos uno nuevo con los encabezados
        if not os.path.exists(EXCEL_FILE):
            df.to_excel(EXCEL_FILE, index=False)
            logger.info("📊 Archivo Excel creado")
        else:
            # Añadir al archivo existente sin sobreescribir
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                # Determinar en qué fila escribir (para evitar sobrescribir datos)
                book = load_workbook(EXCEL_FILE)
                sheet = book.active
                start_row = sheet.max_row + 1
                df.to_excel(writer, index=False, header=False, startrow=start_row-1)
                book.close()
            logger.info("✅ Guardado en Excel correctamente")
        
        return True
    except Exception as e:
        logger.error(f"⚠ Error guardando en Excel: {e}")
        # Guardar en CSV de emergencia si hay error
        try:
            backup_file = f"backup_data_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
            df.to_csv(backup_file, index=False)
            logger.info(f"💾 Respaldo guardado en {backup_file}")
        except:
            pass
        return False

def decodificar_asunto(subject_raw):
    """Decodifica el asunto del correo"""
    try:
        decoded_parts = decode_header(subject_raw)
        subject = ''
        for part, encoding in decoded_parts:
            if isinstance(part, bytes):
                try:
                    subject += part.decode(encoding or 'utf-8', errors='ignore')
                except:
                    subject += part.decode('utf-8', errors='ignore')
            else:
                subject += part
        return subject
    except:
        return subject_raw or ""

def obtener_emails(tabla, status_label, root):
    """Función principal para obtener y procesar emails"""
    global detener_busqueda
    
    try:
        # Actualizar interfaz
        status_label.config(text="🔍 Conectando al servidor IMAP...")
        root.update()
        
        # Conectar al servidor
        logger.info("⏳ Conectando...")
        mail = imaplib.IMAP4_SSL(config.EMAIL_CONFIG[config.EMAIL_PROVIDER]["server"])
        mail.login(config.EMAIL_USER, config.EMAIL_PASS)
        logger.info("✅ Login exitoso")
        mail.select("inbox")

        # Buscar mensajes
        _, mensajes = mail.search(None, "ALL")  # Leemos todos los correos
        correos_filtrados = []
        total_mensajes = len(mensajes[0].split())
        logger.info(f"🔢 Total de correos encontrados: {total_mensajes}")
        
        # Actualizar interfaz
        status_label.config(text=f"🔍 Procesando {total_mensajes} correos...")
        root.update()

        if mensajes and mensajes[0]:
            for num in reversed(mensajes[0].split()):  # Procesar más recientes primero
                # Verificar si la búsqueda debe detenerse
                if detener_busqueda:
                    logger.info("🛑 Búsqueda detenida.")
                    break

                # Obtener mensaje
                _, data = mail.fetch(num, "(RFC822)")
                for respuesta in data:
                    if isinstance(respuesta, tuple):
                        mensaje = email.message_from_bytes(respuesta[1])

                        # Decodificar y normalizar el asunto
                        asunto_raw = mensaje["Subject"]
                        asunto_decodificado = decodificar_asunto(asunto_raw)
                        asunto_normalizado = normalizar_texto(asunto_decodificado).lower()

                        logger.info(f"📬 Asunto detectado: {asunto_normalizado}")

                        # Verificar coincidencias de keywords
                        if not any(keyword in asunto_normalizado for keyword in EMAIL_KEYWORDS):
                            logger.info("⛔ Correo ignorado, no contiene keywords.")
                            continue

                        # Extraer datos
                        remitente = normalizar_texto(mensaje["From"])
                        cuerpo = extraer_cuerpo(mensaje)
                        datos = extraer_datos_estudiante(cuerpo)
                        datos["remitente"] = remitente
                        logger.info(f"🧾 Datos extraídos: {datos}")

                        # Generar ID y carpeta única
                        id_registro, carpeta_id = generar_id_y_carpeta()
                        
                        # Guardar CVs adjuntos
                        rutas_adjuntos = guardar_cv_adjuntos(mensaje, id_registro, datos["nombre"], carpeta_id)
                        
                        # Usar la primera ruta o generar una ruta genérica si no hay adjuntos
                        ruta_cv = rutas_adjuntos[0] if rutas_adjuntos else str(carpeta_id / f"ID_{id_registro}_sin_adjunto.txt")
                        
                        # Guardar en Excel
                        guardado = guardar_en_excel(id_registro, datos, ruta_cv)
                        
                        if guardado:
                            # Añadir a la lista de procesados
                            correos_filtrados.append(datos)
                            
                            # Actualizar tabla en la interfaz
                            tabla.insert("", "end", values=(
                                id_registro, 
                                datos["nombre"], 
                                datos["carrera"], 
                                datos["universidad"], 
                                remitente,
                                rutas_adjuntos[0] if rutas_adjuntos else "Sin adjuntos"
                            ))
                            root.update()

        # Actualizar interfaz al terminar
        status_label.config(text=f"✅ {len(correos_filtrados)} correos procesados.")
        
        # Cerrar conexión
        mail.close()
        mail.logout()

    except Exception as e:
        logger.error(f"❌ Excepción: {e}")
        messagebox.showerror("Error", f"⚠ Error al leer correos: {e}")
        status_label.config(text="❌ Error al procesar correos.")

def iniciar_proceso(tabla, status_label, root):
    """Inicia el proceso en un hilo separado"""
    global detener_busqueda
    detener_busqueda = False  # Asegurarse de que el flag esté en False antes de iniciar
    threading.Thread(target=obtener_emails, args=(tabla, status_label, root), daemon=True).start()

def detener_proceso():
    """Detiene el proceso de búsqueda"""
    global detener_busqueda
    detener_busqueda = True  # Cambiar el flag para detener la búsqueda
    logger.info("🛑 Deteniendo la búsqueda...")

class LectorPanel(tk.Frame):
    """Panel principal de la aplicación"""
    def __init__(self, parent):
        super().__init__(parent, bg="#f5f5f5")
        
        # Etiqueta de estado
        self.status_label = tk.Label(
            self,
            text="🔍 Esperando acción...",
            font=("Segoe UI", 10),
            bg="#f5f5f5"
        )
        self.status_label.pack(pady=(10, 5))

        # Frame para botones en una fila
        botones_frame = tk.Frame(self, bg="#f5f5f5")
        botones_frame.pack(pady=10)

        # Botón para iniciar lectura
        self.btn_iniciar = tk.Button(
            botones_frame,
            text="📥 Leer Correos",
            command=lambda: iniciar_proceso(self.tabla, self.status_label, parent),
            bg="#4CAF50", fg="white",
            font=("Segoe UI", 10),
            padx=10, pady=5
        )
        self.btn_iniciar.pack(side="left", padx=5)

        # Botón para detener búsqueda
        self.btn_detener = tk.Button(
            botones_frame,
            text="⛔ Detener Búsqueda",
            command=detener_proceso,
            bg="#FFA500", fg="white",
            font=("Segoe UI", 10),
            padx=10, pady=5
        )
        self.btn_detener.pack(side="left", padx=5)

        # Botón para salir
        self.btn_finalizar = tk.Button(
            botones_frame,
            text="❌ Finalizar",
            command=parent.quit,
            bg="#F44336", fg="white",
            font=("Segoe UI", 10),
            padx=10, pady=5
        )
        self.btn_finalizar.pack(side="left", padx=5)

        # Frame con scrollbar y tabla
        tabla_frame = tk.Frame(self)
        tabla_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Tabla con más columnas para mostrar IDs
        self.tabla = ttk.Treeview(
            tabla_frame, 
            columns=("ID", "Nombre", "Carrera", "Universidad", "Remitente", "Ruta"), 
            show="headings"
        )
        
        # Configurar columnas
        col_widths = {
            "ID": 50, 
            "Nombre": 200, 
            "Carrera": 200, 
            "Universidad": 200, 
            "Remitente": 200,
            "Ruta": 300
        }
        
        for col, width in col_widths.items():
            self.tabla.heading(col, text=col)
            self.tabla.column(col, width=width, anchor="center")

        # Scrollbar vertical
        scrollbar_y = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=scrollbar_y.set)
        scrollbar_y.pack(side="right", fill="y")
        
        # Scrollbar horizontal
        scrollbar_x = ttk.Scrollbar(tabla_frame, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(xscrollcommand=scrollbar_x.set)
        scrollbar_x.pack(side="bottom", fill="x")

        self.tabla.pack(fill="both", expand=True)

def main():
    """Función principal para iniciar la aplicación"""
    root = tk.Tk()
    root.title("Lector de Correos - Solicitudes de Estadía")
    root.geometry("1200x600")  # Ventana más grande para ver más datos
    
    app = LectorPanel(root)
    app.pack(fill="both", expand=True)
    
    root.mainloop()

if __name__ == "__main__":
    main()